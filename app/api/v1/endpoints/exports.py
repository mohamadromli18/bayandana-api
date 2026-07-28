# app/api/v1/endpoints/exports.py
from fastapi import APIRouter, HTTPException, Depends, Query
from fastapi.responses import StreamingResponse
from datetime import date, datetime
import io

# Excel Imports
import openpyxl
from openpyxl.styles import Font, Alignment

# PDF Imports
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle

# Internal Imports
from app.core.database import supabase
from app.core.security import get_current_org_id

router = APIRouter()

# ==========================================
# 1. ENDPOINT EKSPOR EXCEL
# ==========================================
@router.get("/financial-position/excel")
def export_financial_position_excel(
    as_of_date: date = Query(..., description="Tanggal akhir posisi keuangan (YYYY-MM-DD)"),
    org_id: str = Depends(get_current_org_id)
):
    try:
        # Ambil data akun
        accounts_res = supabase.table("accounts").select("id, account_code, account_name, account_type")\
            .eq("organization_id", org_id)\
            .in_("account_type", ["Asset", "Liability", "Equity"]).execute()
            
        accounts = accounts_res.data
        if not accounts:
            raise HTTPException(status_code=404, detail="Tidak ada data akun.")

        # Ambil data mutasi
        mutations_res = supabase.table("journal_details").select(
            "account_id, debit, credit, journals!inner(transaction_date)"
        ).eq("journals.organization_id", org_id)\
         .lte("journals.transaction_date", as_of_date.isoformat()).execute()
         
        account_balances = {acc["id"]: 0.0 for acc in accounts}
        account_types = {acc["id"]: acc["account_type"] for acc in accounts}
        
        for mut in mutations_res.data:
            acc_id = mut["account_id"]
            if acc_id in account_balances:
                debit = float(mut["debit"])
                credit = float(mut["credit"])
                if account_types[acc_id] == "Asset":
                    account_balances[acc_id] += (debit - credit)
                else:
                    account_balances[acc_id] += (credit - debit)

        # Inisialisasi Workbook Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Posisi Keuangan"

        bold_font = Font(bold=True)

        # Tulis Header Dokumen
        ws.append(["LAPORAN POSISI KEUANGAN (NERACA)"])
        ws["A1"].font = bold_font
        ws.append([f"Per Tanggal: {as_of_date.isoformat()}"])
        ws.append([])

        # Header Tabel
        headers = ["Kode Akun", "Nama Akun", "Saldo (Rp)"]
        ws.append(headers)
        for col in range(1, 4):
            ws.cell(row=4, column=col).font = bold_font

        # Data Aset
        ws.append(["ASET"])
        ws.cell(row=ws.max_row, column=1).font = bold_font
        total_assets = 0.0
        for acc in accounts:
            if acc["account_type"] == "Asset":
                balance = account_balances[acc["id"]]
                ws.append([acc["account_code"], acc["account_name"], balance])
                total_assets += balance
        ws.append(["Total Aset", "", total_assets])
        ws.cell(row=ws.max_row, column=1).font = bold_font
        ws.cell(row=ws.max_row, column=3).font = bold_font
        ws.append([])

        # Data Kewajiban
        ws.append(["KEWAJIBAN"])
        ws.cell(row=ws.max_row, column=1).font = bold_font
        total_liabilities = 0.0
        for acc in accounts:
            if acc["account_type"] == "Liability":
                balance = account_balances[acc["id"]]
                ws.append([acc["account_code"], acc["account_name"], balance])
                total_liabilities += balance
        ws.append(["Total Kewajiban", "", total_liabilities])
        ws.cell(row=ws.max_row, column=1).font = bold_font
        ws.cell(row=ws.max_row, column=3).font = bold_font
        ws.append([])

        # Data Saldo Dana (Ekuitas)
        ws.append(["SALDO DANA (EKUITAS)"])
        ws.cell(row=ws.max_row, column=1).font = bold_font
        total_equities = 0.0
        for acc in accounts:
            if acc["account_type"] == "Equity":
                balance = account_balances[acc["id"]]
                ws.append([acc["account_code"], acc["account_name"], balance])
                total_equities += balance
        ws.append(["Total Saldo Dana", "", total_equities])
        ws.cell(row=ws.max_row, column=1).font = bold_font
        ws.cell(row=ws.max_row, column=3).font = bold_font
        ws.append([])

        # Total Kewajiban & Saldo Dana
        ws.append(["Total Kewajiban & Saldo Dana", "", total_liabilities + total_equities])
        ws.cell(row=ws.max_row, column=1).font = bold_font
        ws.cell(row=ws.max_row, column=3).font = bold_font

        # Lebarkan kolom
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 20

        # Simpan ke Stream
        stream = io.BytesIO()
        wb.save(stream)
        stream.seek(0)

        filename = f"Neraca_{as_of_date.isoformat()}.xlsx"
        return StreamingResponse(
            stream, 
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", 
            headers={'Content-Disposition': f'attachment; filename="{filename}"'}
        )

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ==========================================
# 2. ENDPOINT EKSPOR PDF
# ==========================================
@router.get("/financial-position/pdf")
def export_financial_position_pdf(
    as_of_date: date = Query(..., description="Tanggal akhir posisi keuangan (YYYY-MM-DD)"),
    org_id: str = Depends(get_current_org_id)
):
    try:
        # --- LOGIKA QUERY DATA (SAMA DENGAN EXCEL) ---
        accounts_res = supabase.table("accounts").select("id, account_code, account_name, account_type")\
            .eq("organization_id", org_id)\
            .in_("account_type", ["Asset", "Liability", "Equity"]).execute()
            
        accounts = accounts_res.data
        if not accounts:
            raise HTTPException(status_code=404, detail="Tidak ada data akun.")

        mutations_res = supabase.table("journal_details").select(
            "account_id, debit, credit, journals!inner(transaction_date)"
        ).eq("journals.organization_id", org_id)\
         .lte("journals.transaction_date", as_of_date.isoformat()).execute()
         
        account_balances = {acc["id"]: 0.0 for acc in accounts}
        account_types = {acc["id"]: acc["account_type"] for acc in accounts}
        
        for mut in mutations_res.data:
            acc_id = mut["account_id"]
            if acc_id in account_balances:
                debit = float(mut["debit"])
                credit = float(mut["credit"])
                if account_types[acc_id] == "Asset":
                    account_balances[acc_id] += (debit - credit)
                else:
                    account_balances[acc_id] += (credit - debit)

        # --- PEMBUATAN PDF DENGAN REPORTLAB ---
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=40, leftMargin=40, topMargin=40, bottomMargin=40)
        
        elements = []
        styles = getSampleStyleSheet()
        
        # Styles
        title_style = ParagraphStyle(name="TitleCenter", parent=styles['Heading1'], alignment=1, spaceAfter=6)
        subtitle_style = ParagraphStyle(name="SubtitleCenter", parent=styles['Normal'], alignment=1, spaceAfter=20)
        
        # Header Dokumen PDF
        # Anda dapat mengganti "NAMA ORGANISASI" dengan query nama organisasi jika diperlukan
        elements.append(Paragraph("LAPORAN POSISI KEUANGAN (NERACA)", title_style))
        elements.append(Paragraph(f"Per Tanggal: {as_of_date.isoformat()}", subtitle_style))
        elements.append(Spacer(1, 12))
        
        # Persiapan Data Tabel
        table_data = [["Kode Akun", "Nama Akun", "Saldo (Rp)"]]
        
        # Data ASET
        table_data.append(["", "ASET", ""])
        total_assets = 0.0
        for acc in accounts:
            if acc["account_type"] == "Asset":
                balance = account_balances[acc["id"]]
                table_data.append([acc["account_code"], acc["account_name"], f"{balance:,.2f}"])
                total_assets += balance
        table_data.append(["", "Total Aset", f"{total_assets:,.2f}"])
        table_data.append(["", "", ""]) # Spacer
        
        # Data KEWAJIBAN
        table_data.append(["", "KEWAJIBAN", ""])
        total_liabilities = 0.0
        for acc in accounts:
            if acc["account_type"] == "Liability":
                balance = account_balances[acc["id"]]
                table_data.append([acc["account_code"], acc["account_name"], f"{balance:,.2f}"])
                total_liabilities += balance
        table_data.append(["", "Total Kewajiban", f"{total_liabilities:,.2f}"])
        table_data.append(["", "", ""]) # Spacer
        
        # Data SALDO DANA (EKUITAS)
        table_data.append(["", "SALDO DANA (EKUITAS)", ""])
        total_equities = 0.0
        for acc in accounts:
            if acc["account_type"] == "Equity":
                balance = account_balances[acc["id"]]
                table_data.append([acc["account_code"], acc["account_name"], f"{balance:,.2f}"])
                total_equities += balance
        table_data.append(["", "Total Saldo Dana", f"{total_equities:,.2f}"])
        table_data.append(["", "", ""]) # Spacer
        
        # Total Kewajiban & Ekuitas
        total_liab_eq = total_liabilities + total_equities
        table_data.append(["", "Total Kewajiban & Saldo Dana", f"{total_liab_eq:,.2f}"])
        
        # Setting Kolom Tabel ReportLab
        t = Table(table_data, colWidths=[80, 280, 140])
        
        # Styling Tabel PDF
        t.setStyle(TableStyle([
            # Header Style
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2c3e50')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
            
            # Alignments & Borders
            ('ALIGN', (0, 1), (0, -1), 'CENTER'),
            ('ALIGN', (1, 1), (1, -1), 'LEFT'),
            ('ALIGN', (2, 1), (2, -1), 'RIGHT'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            
            # Sub-headers (ASET, KEWAJIBAN, SALDO DANA)
            ('FONTNAME', (1, 1), (1, 1), 'Helvetica-Bold'),
            ('BACKGROUND', (0, 1), (-1, 1), colors.HexColor('#ecf0f1')),
            
            # Total Penyorotan (Menyesuaikan index baris tergantung jumlah akun, lebih aman pakai bold manual atau format spesifik)
            # Karena indeks baris total bersifat dinamis, kita buat Helvetica rata-rata, 
            # Jika ingin menebalkan total secara spesifik di ReportLab lebih baik memisahkan logic indeks.
            # Namun di sini grid sudah cukup rapi.
        ]))
        
        elements.append(t)
        
        # Render PDF
        doc.build(elements)
        buffer.seek(0)
        
        filename = f"Neraca_{as_of_date.isoformat()}.pdf"
        return StreamingResponse(
            buffer, 
            media_type="application/pdf", 
            headers={'Content-Disposition': f'attachment; filename="{filename}"'}
        )

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))